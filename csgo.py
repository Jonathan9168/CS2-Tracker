import time
import config
import requests
import pandas as pd
from datetime import datetime
from bs4 import BeautifulSoup
from urllib.parse import quote
from openpyxl import load_workbook


def get_current_item_value_steam(name, max_retries=1, ttw=3):
    """

    Retrieves CSGO item price from Steam endpoint

    name: item name
    max_retries: maximum number of times to attempt to fetch price data for item if the request was unsuccessful
    ttw (time to wait): to prevent error 429, add wait time between requests
    response: reply from Steam endpoint

    response format (successful):
     {
        "success": true,
        "lowest_price": "£4.22",
        "volume": "250",
        "median_price": "£4.37"
    }

    response format (unsuccessful):
    {
        "success": false
    }

    total script time ~ ttw * no. unique items (ttw= 0 can be used for < 18 unique items consistently before error 429)
    :returns: current Steam market listing price (successful) or False (unsuccessful)

    """

    response_data = None

    for _ in range(max_retries):
        try:
            link = f"https://steamcommunity.com/market/priceoverview/?currency=2&appid=730&market_hash_name={quote(name)}"
            response = requests.get(link)
            time.sleep(ttw)

            if response.ok:

                response_data = response.json()

                if response_data["success"] is True:
                    return float(response_data['lowest_price'][1:])
                else:
                    continue

        except KeyError as e:
            if "median_price" in response_data:
                return float(response_data['median_price'][1:])

            print(response_data)
            print("An error occurred:", e)

    return False


def get_current_item_value_cs_trader(name):
    """

    Retrieves CSGO item price from CSGO Trader (updated every 8 hours)

    cs_trader_json structure:

    "<item_name>": {
    "steam": {
      "last_24h": 0.03,
      "last_7d": 0.03,
      "last_30d": 0.03,
      "last_90d": 0.03
    },
    "lootfarm": null,
    "csgotm": "0.006",
    "skinport": { "suggested_price": 0.03, "starting_at": 0.02 },
    "csgoempire": 0.01,
    "swapgg": 0.03,
    "csgoexo": null,
    "cstrade": null,
    "skinwallet": null,
    "buff163": {
      "starting_at": { "price": null },
      "highest_order": { "price": null }
        }
    }

    """

    try:
        price = 0

        if option == "b":
            price = cs_trader_json[name]["steam"]["last_24h"]
        elif option == "c":
            price = cs_trader_json[name]["steam"]["last_7d"]

        value = float(price) * float(conversion_rate)

        # Steam price floor value
        if value < 0.03:
            return 0.03
        return value

    except KeyError as e:
        print("An error occurred:", e)


def percentage_change(old_value, new_value):
    """

    Calculates percentage change between two values

    old_value: previous value of some variable
    new_value: updated value of some variable

    :returns: percentage difference between old and new value (not multiplied by 100 as Excel percentage format does this)

    """

    # use 0.01 as a placeholder to avoid division by 0 error
    if old_value == 0:
        old_value = 0.01

    return (new_value - old_value) / old_value


def calculate_expected_profit():
    """
    Calculates the expected profit in cell L1 given by the formula: =SUMIFS(H:H, I:I, "N/A") * 0.85

    We take the sum of values in column H (Current - Purchase [DIFF])(difference in current market value and purchase price)
    where the corresponding value in column I (Sold Price) is "N/A",
    this means that only items that have not been sold are factored into the calculation.

    Finally, the sum is multiplied by 0.85 to account for Steam selling fees (5% Steam + 10% game fee [CSGO])

    :returns: The expected profit if the items are sold at Steam market value

    """
    return 0.85 * sum([float(current_value_steam.value) - float(purchase_price.value) for
                       purchase_price, current_value_steam, sold_price in zip(ws['E'][1:], ws['F'][1:], ws['I'][1:]) if
                       sold_price.value == "N/A"])


def update_dataframe():
    """

    Updates dataframe clone of Excel file with new item values

    item: item value as seen in spreadsheet column 'B'
    condition: item value as seen in spreadsheet column 'C'
    item_name: item name as seen on the Steam marketplace

    item_name format (skin): "<Weapon> | (<Condition>)"
    item_name format (other): "<Item>"

    """

    for index, row in df.iterrows():

        item, condition = row['Item'], row['Condition']
        item_name = f"{item.strip()}" if pd.isnull(condition) else f"{item.strip()} ({condition.strip()})"

        print(f"[{index}] {item_name}")

        # Skin has been processed before so avoid calling API by retrieving info from cache
        if item_name in items_processed and item_name in percentage_changes:
            update_dataframe_on_success(index, item_name)
            continue

        # Skin hasn't been processed yet
        if option == "a":
            current_value = get_current_item_value_steam(name=item_name)
        else:
            current_value = get_current_item_value_cs_trader(name=item_name)

        if current_value is False:  # Invalid item due to typo in file or JSON response failure, don't update item and write "n"
            update_dataframe_failure(index)
        else:  # Skin is valid: update processed items and percentage change cache, update item's;  value percentage change, current value and update item value modified status
            update_cache(item_name, current_value, index)
            update_dataframe_on_success(index, item_name)


def update_cache(item_name, current_value, index):
    """

    Updates processed skins caches for: item current value and percentage change for item value, minimizes calls to
    Steam endpoint

    """

    items_processed[item_name] = current_value
    percentage_changes[item_name] = percentage_change(df.at[index, 'Current Value [Steam]'], current_value)


def update_dataframe_on_success(index, item_name):
    """

    Updates 'Current Value', 'Current Value % Change', and 'Current Value Updated' columns in data frame
    if item has been processed before or on successful response from Steam endpoint

    """

    df.at[index, 'Current Value % Change'] = percentage_changes[item_name]
    df.at[index, 'Current Value [Steam]'] = items_processed[item_name]
    df.at[index, 'Current Value Updated'] = "y"


def update_dataframe_failure(index):
    """Updates 'Current Value Updated' column when response from Steam endpoint is unsuccessful"""

    df.at[index, 'Current Value Updated'] = "n"


def dataframe_to_excel():
    """Update Excel file with updated values from dataframe"""

    # Write the updated "Current Value," "Current Value % Change," and "Current Value Updated" columns to the Excel file
    for index, (current_value, current_value_change, current_value_updated) in enumerate(
            zip(df['Current Value [Steam]'], df['Current Value % Change'], df['Current Value Updated']), start=2):
        cell_current_value = f'F{index}'  # Assuming the "Current Value [Steam]" column is column F
        cell_current_value_change = f'G{index}'  # Assuming the "Current Value % Change" column is column G
        cell_current_value_updated = f'J{index}'  # Assuming the "Current Value Updated" column is column J

        ws[cell_current_value] = current_value
        ws[cell_current_value_change] = current_value_change
        ws[cell_current_value_updated] = current_value_updated

    update_expected_percentage_change()
    update_time_modified()


def update_expected_percentage_change():
    """Updates cell M1 with expected profit percentage change"""

    new_expected_profit = calculate_expected_profit()
    ws['M1'] = percentage_change(old_expected_profit, new_expected_profit)


def update_time_modified():
    """ Updates cell L3 with the current date and time"""

    current_time = datetime.now().strftime("%d/%m/%Y at %H:%M")
    ws['L3'] = current_time  # Writing update time into cell L3


def save_excel():
    """Saves the updated Excel file with the original formatting to specified directories"""

    wb.save(file_path_local)

    if file_path_desktop is not None:
        wb.save(file_path_desktop)


def get_conversion_rate():
    """Retrieves USD -> GBP conversion rate"""

    url = "https://www.xe.com/currencyconverter/convert/?Amount=1&From=USD&To=GBP"
    res = requests.get(url)
    soup = BeautifulSoup(res.content, 'html.parser')

    return soup.find('p', class_='result__BigRate-sc-1bsijpp-1 iGrAod').text[:-14]


def main_menu():
    """Option Menu"""

    print("\n[A] Update current values using live Steam Prices")
    print("[B] Update current values using CSGO Trader 24hr Avg [Updated every 8 hours]")
    print("[C] Update current values using CSGO Trader 7day Avg [Updated every 8 hours]")
    choice = input("Select an option: ").strip().lower()
    return choice


if __name__ == "__main__":

    option = main_menu()
    valid_options = {"a", "b", "c"}

    if option not in valid_options:
        print("invalid option")
        quit(0)

    if option == "b" or option == "c":
        conversion_rate = get_conversion_rate()
        cs_trader_json = requests.get("https://prices.csgotrader.app/latest/prices_v6.json").json()

    # Load Excel spreadsheet into workbook and Pandas dataframe
    file_path_local = config.file_path_local
    file_path_desktop = config.file_path_desktop

    df = pd.read_excel(file_path_local)
    wb = load_workbook(file_path_local)
    ws = wb.active

    # Processing caches
    items_processed, percentage_changes = {}, {}

    # Previous expected profit to be used in % change calculation later
    old_expected_profit = calculate_expected_profit()

    update_dataframe()
    dataframe_to_excel()
    save_excel()
