import re
import time
import config
from copy import copy
from datetime import datetime
from selenium import webdriver
from openpyxl.styles import Font
import chromedriver_autoinstaller
from openpyxl import load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

chromedriver_autoinstaller.install()


def generate_driver():
    """Generating chromedriver instance"""

    chromeOption = Options()
    chromeOption.add_argument("--headless")
    chromeOption.add_argument("--window-size=1600,1200")
    chromeOption.add_argument("--mute-audio")
    chromeOption.add_argument("--log-level=3")
    chromeOption.add_argument("--silent")
    chromeOption.add_argument("--disable-blink-features=AutomationControlled")
    chromeOption.add_experimental_option("excludeSwitches", ['enable-automation'])

    if config.chrome_driver_executable_path is None:
        driver = webdriver.Chrome(options=chromeOption)
    else:
        driver = webdriver.Chrome(
            options=chromeOption,
            service=Service(executable_path=config.chrome_driver_executable_path)
        )

    return driver


def scrape_inventory():
    """Scrapes only marketable items from a specified Steam inventory

    flag: determines whether to stop scraping based on an empty inventory slot or disabled forward pagination button
    processed_count: keeps track of how many items have been processed so seen items are not reprocessed
    max_row: alternate max_row (starts from row 1) as opposed to ws.max_row (starts at 3 due to 'Last Price Check' cell),
    allows population from row 1 without leaving two space gap in the spreadsheet

    """

    flag, processed_count, max_row = True, 0, 1

    # generate chromedriver and accept cookies
    driver = generate_driver()
    driver.get(base_url)
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, 'acceptAllButton'))).click()

    print("Scraping Items...\n")

    while flag:

        # find item holder elements
        items = driver.find_elements(By.CLASS_NAME, 'itemHolder')

        if len(items) == 0:
            print("The inventory is private or unavailable.")
            quit(0)

        for i, item in enumerate(items[processed_count:]):

            if 'disabled' not in item.get_attribute('class'):  # check if item slot is empty

                item.click()
                rhs = driver.find_element(By.CLASS_NAME, 'inventory_page_right')

                # for some reason, the most recent item name and tag elements alternate when clicking a new item
                if i % 2 == 1:
                    item_name_element = rhs.find_elements(By.CLASS_NAME, 'hover_item_name')[1]
                    item_tag_element = rhs.find_elements(By.CSS_SELECTOR, 'span.item_desc_descriptors')[1]
                else:
                    item_name_element = rhs.find_elements(By.CLASS_NAME, 'hover_item_name')[0]
                    item_tag_element = rhs.find_elements(By.CSS_SELECTOR, 'span.item_desc_descriptors')[0]

                item_name = item_name_element.text.strip()
                item_tag = item_tag_element.text.strip()

                print(f"[{processed_count + 1}] {item_name:^50} -> [{item_tag}]")
                processed_count += 1

                # check if the item is marketable (has value)
                if item_tag.split(",")[-1].strip() == "Marketable":
                    add_item_to_excel(item_name, item_tag, max_row)
                    max_row += 1

            else:
                flag = False
                break

            # if last item on the page has been processed (25 per page) find the 'next' button and check if it's disabled
            if (i + 1) % 25 == 0:

                next_button = driver.find_element(By.ID, 'pagebtn_next')

                if 'disabled' in next_button.get_attribute('class'):
                    flag = False
                    break

                next_button.click()
                time.sleep(.5)

    driver.quit()


def add_item_to_excel(item_name, item_tag, max_row):
    """Adds inventory item into spreadsheet as a new row with necessary information

    condition: checks if item has a condition value e.g. weapon, if not then leave blank
    item_tag: contains information relevant to the item, some examples below

    [Pistol, P250, The Breakout Collection, Normal, Restricted, Factory New, Tradable, Marketable]
    [Graffiti, Normal, Base Grade, Jungle Green, Tradable, Marketable]
    [Music Kit, Normal, High Grade, Not Tradable, Not Marketable]
    [Collectible, Normal, Extraordinary, Not Tradable, Not Marketable]
    [Container, The Clutch Collection, Normal, Base Grade, Tradable, Marketable]

    hence, we focus on index 0 using 'conditional_items_filter' to check if the item is a weapon before subscripting for
    item condition

    """

    date = f'{datetime.now().strftime("%d/%m/%Y")}'
    item = item_name

    split_values = item_tag.split(',')

    # special Glove/Knife condition cases
    wear_index = (
        6 if split_values[0] == "Knife" and len(split_values) >= 9 else
        3 if split_values[0] == "Gloves" else
        4 if split_values[0] == "Knife" else
        5
    )

    condition = f"{split_values[wear_index] if split_values[0] in conditional_items_filter else ''}"
    purchase_platform = "Steam"
    purchase_price = 0.03
    current_value = 0.03
    percentage_change = 0.00
    price_difference = 0.00
    sold_price = "N/A"
    updated = "n"

    data = [date, item, condition, purchase_platform, purchase_price, current_value, percentage_change,
            price_difference, sold_price, updated]

    # create a new row
    new_row = max_row + 1

    # create new cell in new row and apply cell styling from header row
    for col_index, cell in enumerate(row_styles, start=1):
        new_cell = ws.cell(row=new_row, column=col_index, value=None)  # Create an empty cell
        new_cell._style = copy(row_styles[col_index - 1]._style)  # Copy the styling

    # populate cells with scraped item values
    for j, new_cell in enumerate(data):
        ws.cell(row=new_row, column=j + 1, value=data[j])

    item_tag_list = item_tag.split(",")
    item_tag_list_stripped = [s.strip() for s in item_tag_list]

    name_colour = get_color(item_tag_list_stripped)

    if name_colour is not None:
        name_colour_aRGB = RGB_Hex_To_aRGB_Hex(name_colour)
        name_cell = ws.cell(row=new_row, column=2)  # Assuming item name is in the second column

        font = Font(color=name_colour_aRGB, bold="yes", name="Open Sans", sz=10)
        name_cell.font = font


def apply_difference_formula():
    """Applies price difference formula to column H"""

    # applying the formula from row 2 onward
    for row_index in range(2, ws.max_row + 1):
        formula_cell = f'H{row_index}'  # The cell in column H for the formula
        formula = f'=F{row_index}-E{row_index}'  # =Fx-Ex
        ws[formula_cell] = formula


def save_excel():
    """Saves the updated Excel file with the original formatting to specified directories"""

    wb.save(file_path_local)

    if file_path_desktop is not None:
        wb.save(file_path_desktop)


def get_color(tag_list):
    """Retrieves item rarity colour"""

    for item in tag_list:
        if item.strip() in item_rarities:
            return item_rarities[item]
    return None


def RGB_Hex_To_aRGB_Hex(RGB_Hex):
    """Converts RGB Hex to aRGB format"""

    aRGB_Hex = 'FF' + RGB_Hex[1:]
    return aRGB_Hex


if __name__ == "__main__":

    # user Steam inventory URL input combined with #730 for CS2
    pattern = r'https://steamcommunity\.com/(id|profiles)/[\w-]+/inventory/'
    base_url = input("Steam Inventory URL: ")

    if not re.match(pattern, base_url):
        print("Invalid URL")
        quit(0)

    base_url += "#730"

    # load Excel spreadsheet into workbook
    base_path = config.base_path
    file_path_local = config.file_path_local
    file_path_desktop = config.file_path_desktop

    wb = load_workbook(base_path)
    ws = wb.active

    # get first row of cells from column A-J so cell in column can inherit styling
    row_styles = ws[1][:10]

    # valid categories to help filter for wear-able items
    conditional_items_filter = {"Rifle", "SMG", "Shotgun", "Pistol", "Sniper Rifle", "Knife", "Machinegun", "Gloves"}

    # item rarity colour codes
    item_rarities = {
        'Consumer Grade': '#B0C3D9',
        'Industrial Grade': '#5E98D9',
        'Mil-Spec Grade': '#4B69FF',
        'Restricted': '#8847FF',
        'Classified': '#D32CE6',
        'Covert': '#EB4B4B',
        'Contraband': '#E4AE33',
        'Clandestine': '#E4AE33',
        'UNNAMED': '#ADE55C',
        'Base Grade': '#B0C3D9',
        'Medium Grade': '#5E98D9',
        'High Grade': '#4B69FF',
        'Remarkable': '#8847FF',
        'Exotic': '#D32CE6',
        'Distinguished': '#4B69FF',
        'Exceptional': '#8847FF',
        'Superior': '#D32CE6',
        'Master': '#EB4B4B',
    }

    scrape_inventory()
    apply_difference_formula()
    save_excel()
